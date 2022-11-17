import xlwings as xw

# filepath=r'g:\Python Scripts\test.xlsx'
# wb=app.books.open(filepath)
# wb.save()
# wb.close()
# app.quit()

# import xlwings as xw
# app=xw.App(visible=False,add_book=False)
# wb=app.books.add()
# wb.save(r'd:\test.xlsm')
# wb.close()
# app.quit()

import re
import os

import time

import win32com

# from win32com.client import Dispatch

#可以修改文件类型，但是需要吧当前打开的excel关掉，否则会直接干掉当前打开的excel进程
def xls_xlsx(path):

    w = win32com.client.Dispatch('Excel.Application')

    w.Visible = 0

    w.DisplayAlerts = 0

    wb = w.Workbooks.Open(path)
    newpath = os.path.splitext(path)[0] + "_Test.xlsm"
    wb.SaveAs(newpath,FileFormat = 52)

    # doc.Close() 开启则会删掉原来的dxls

    w.Quit()# 退出

    # return newpath

# allpath = os.getcwd()
#
# print(allpath)
# ExcelPath = r"E:\GitHub\MiniTool\DataSource\CHT_SWV_GWM_P0102_2S_EOL_Test_Result.xlsm"
# xls_xlsx(ExcelPath)

import pandas as pd

# df = pd.DataFrame({'First' : [5, 2, 0, 10, 4],
# #                    'Second' : [9, 8, 21, 3, 8]})
# #
# # writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
# #
# # df.to_excel(writer, sheet_name='Sheet1')
# #
# # workbook  = writer.book
# # workbook.filename = 'test.xlsm'
# # workbook.add_vba_project('Enable_Disable.bas')
# #
# # writer.save()

import openpyxl
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

def Xlsx2Xlsm(path):


    wb = openpyxl.load_workbook(path, keep_vba=True)
    newpath = os.path.splitext(path)[0] + "_Test.xlsm"
    wb.save(newpath)

# Xlsx2Xlsm(ExcelPath)
#
# wb = Workbook()
# dest_filename = 'empty_book.xlsx'
# wb = openpyxl.load_workbook(dest_filename,keep_vba=True)
# wb.save('empty_book.xlsm')
# ws1 = wb.active
# ws1.title = "range names"
#
# for row in range(1, 40):
#  ws1.append(range(600))
#
# ws2 = wb.create_sheet(title="Pi")
#
# ws2['F5'] = 3.14
#
# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#  for col in range(27, 54):
# 	 _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['AA10'].value)

# wb.save(filename = dest_filename)



class easyExcel:
    """A utility to make it easier to get at Excel.    Remembering
    to save the data is your problem, as is    error handling.
    Operates on one workbook at a time."""

    def __init__(self, filename=None):  # 打开文件或者新建文件（如果不存在的话）
        self.xlApp = win32com.client.Dispatch('Excel.Application')
        if filename:
            self.filename = filename
            self.xlBook = self.xlApp.Workbooks.Open(filename)
        else:
            self.xlBook = self.xlApp.Workbooks.Add()
            self.filename = ''

    def save(self, newfilename=None):  # 保存文件
        if newfilename:
            self.filename = newfilename
            self.xlBook.SaveAs(newfilename)
        else:
            self.xlBook.Save()

    def close(self):  # 关闭文件
        self.xlBook.Close(SaveChanges=0)
        del self.xlApp

    def getCell(self, sheet, row, col):  # 获取单元格的数据
        "Get value of one cell"
        sht = self.xlBook.Worksheets(sheet)
        return sht.Cells(row, col).Value

    def setCell(self, sheet, row, col, value):  # 设置单元格的数据
        "set value of one cell"
        sht = self.xlBook.Worksheets(sheet)
        sht.Cells(row, col).Value = value

    def setCellformat(self, sheet, row, col):  # 设置单元格的数据
        "set value of one cell"
        sht = self.xlBook.Worksheets(sheet)
        sht.Cells(row, col).Font.Size = 15  # 字体大小
        sht.Cells(row, col).Font.Bold = True  # 是否黑体
        sht.Cells(row, col).Name = "Arial"  # 字体类型
        sht.Cells(row, col).Interior.ColorIndex = 3  # 表格背景
        # sht.Range("A1").Borders.LineStyle = xlDouble
        sht.Cells(row, col).BorderAround(1, 4)  # 表格边框
        sht.Rows(3).RowHeight = 30  # 行高
        sht.Cells(row, col).HorizontalAlignment = -4131  # 水平居中xlCenter
        sht.Cells(row, col).VerticalAlignment = -4160  #

    def deleteRow(self, sheet, row):
        sht = self.xlBook.Worksheets(sheet)
        sht.Rows(row).Delete()  # 删除行
        sht.Columns(row).Delete()  # 删除列

    def getRange(self, sheet, row1, col1, row2, col2):  # 获得一块区域的数据，返回为一个二维元组
        "return a 2d array (i.e. tuple of tuples)"
        sht = self.xlBook.Worksheets(sheet)
        return sht.Range(sht.Cells(row1, col1), sht.Cells(row2, col2)).Value

    def addPicture(self, sheet, pictureName, Left, Top, Width, Height):  # 插入图片
        "Insert a picture in sheet"
        sht = self.xlBook.Worksheets(sheet)
        sht.Shapes.AddPicture(pictureName, 1, 1, Left, Top, Width, Height)

    def cpSheet(self, before):  # 复制工作表
        "copy sheet"
        shts = self.xlBook.Worksheets
        shts(1).Copy(None, shts(1))

    def inserRow(self, sheet, row):
        sht = self.xlBook.Worksheets(sheet)
        sht.Rows(row).Insert(1)


ExcelPath = r"E:\GitHub\MiniTool\DataSource\CHT_SWV_GWM_P0102_2S_EOL_Test_Result.xlsm"
#


def ReplaceCellValue(CellValue,Df_LoopUp):
    CellValue_list = CellValue.split('\n')
    returnValue_List = []
    # reg =
    for i,cell in enumerate(CellValue_list):
        if cell.strip() in Df_LoopUp.index:
            returnValue_List.append(Df_LoopUp.loc[cell.strip(), "CBID"])
        elif cell.strip() == "":
            pass
        else:
            returnValue_List.append(cell.strip())
    # print(returnValue_List)
    # print("*"*30)
    returnValue = "\n".join(returnValue_List)
    # returnValue = returnValue.replace(" ","").strip()
    return returnValue
    # return CellValue_list
#1. 读取spec的
def ReadSpec_ReplaceDoorsID(Spec,Df_LoopUp):
    #获取最后一个sheet的名字
    Df_AllSheet = pd.read_excel(Spec, sheet_name=None)
    LastSheet_Name = list(Df_AllSheet)[-1]
    # print(LastSheet_Name)
    Df_spec = pd.read_excel(Spec, LastSheet_Name, dtype='str',header= 6)

    ColumnsList = ["Test Case Name", "_VerifiesDOORSRequirements_SW"]
    Df_spec = Df_spec[ColumnsList]
    Df_spec = Df_spec.fillna("")

    Df_spec['_VerifiesDOORSRequirements_SW'] = Df_spec['_VerifiesDOORSRequirements_SW'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    # print(Df_PTC_Spec)
    Df_spec.set_index("Test Case Name", inplace=True)
    #获取行数
    RowSize = len(Df_spec.index)

    # print(RowSize)
    return Df_spec,LastSheet_Name,RowSize
    #NA


def ReadLoopUp(LoopUP):
    ColumnsList = ["CBID", "DoorsID"]
    Df_LoopUp = pd.read_csv(LoopUP, names = ColumnsList, header = None,sep = "\t",dtype= 'str')

    Df_LoopUp.set_index("DoorsID", inplace=True)
    # print(Df_LoopUp)
    # print(Df_LoopUp.loc["DES_SA2FC_SW_REQ_BOOT_585","CBID"])

    return Df_LoopUp



# '''
# # 可以重新建立一个独立的进程
def RepaceDoorsID_SaveMacroEcel(w,Spec,Df_spec,LastSheet_Name,RowSize):

    # w = win32com.client.DispatchEx('Excel.Application')
    # w.Visible = 0
    # w.DisplayAlerts = 0
    # # print(os.path.abspath(Spec))
    wb = w.Workbooks.Open(os.path.abspath(Spec))

    try:
        sht = wb.Worksheets(LastSheet_Name)
        for i in range(8,RowSize + 8):
            if sht.Cells(i, "B").Value in Df_spec.index:
                # print(sht.Cells(i, "B").Value)
                sht.Cells(i, "M").Value = Df_spec.loc[sht.Cells(i, "B").Value,"_VerifiesDOORSRequirements_SW"]
    except:
        pass
    wb.Close(SaveChanges=1)


def SaveErrorDefinitionFromFltMonr_Configurator(FltMonr_Configurator,ErrorDefinition):
    """
    使用 win32com.client.DispatchEx('Excel.Application') 将从FltMonr_Configurator 生成ErrorDefinition文件
    :param FltMonr_Configurator:
    :param ErrorDefinition:
    :return: NONE
    """
    ExcelAPP = win32com.client.DispatchEx('Excel.Application')
    ExcelAPP.Visible = 0
    ExcelAPP.DisplayAlerts = 0
    wb = ExcelAPP.Workbooks.Open(os.path.abspath(FltMonr_Configurator))

    for sheetObj in wb.Worksheets:
        # print(sheetObj.Name)
        if sheetObj.Name == "ACCT Autoliv Faults":
            sheetObj.Name = "Errors"
        else:
            wb.Worksheets(sheetObj.Name).Delete()
    wb.SaveAs(os.path.abspath(ErrorDefinition))
    wb.Close(SaveChanges=0)
    ExcelAPP.Quit()

def SaveErrorDefinitionFromFltMonr_Configurator2(FltMonr_Configurator,ErrorDefinition):
    ExcelAPP = win32com.client.DispatchEx('Excel.Application')
    ExcelAPP.Visible = 0
    ExcelAPP.DisplayAlerts = 0
    wb = ExcelAPP.Workbooks.Open(os.path.abspath(FltMonr_Configurator))

    wb2 = ExcelAPP.Workbooks.Open(os.path.abspath(ErrorDefinition))

    for sheetObj in wb.Worksheets:
        # print(sheetObj.Name)
        if sheetObj.Name == "ACCT Autoliv Faults":
            sheetObj.Name = "Errors2"
            wb.Worksheets("Errors2").Copy(wb2.Worksheets("Errors"))
        # else:
        #     wb.Worksheets(sheetObj.Name).Delete()
    # wb2.SaveAs(ErrorDefinition)
    wb2.Close(SaveChanges=1)
    wb.Close(SaveChanges=0)
    ExcelAPP.Quit()
    #
    # w.Quit()  # 退出
# '''
if __name__ == '__main__':
    ExcelPath = r"C:\Project\Geely_GEEA2_HX11\SW_Release\P23\BuildGenerationReport\FltMonr_Configurator_SC3.xlsm"
    newPath = r"C:\Project\Geely_GEEA2_HX11\SW_Release\P23\BuildGenerationReport\error_definition_Geely_GEEA2_HX11_P25.xlsm"
    SaveErrorDefinitionFromFltMonr_Configurator(ExcelPath,newPath)
    # wb = openpyxl.load_workbook(ExcelPath, keep_vba=True,data_only= True,keep_links=False)
    # wb2 = openpyxl.Workbook()
    # # copysheet = wb["ACCT Autoliv Faults"]
    # copysheet = wb2.copy_worksheet(wb["ACCT Autoliv Faults"])
    # # sheetNames = wb.sheetnames;
    # # for sheetName in sheetNames:
    # #     ws = wb[sheetName]
    # #     if sheetName == "ACCT Autoliv Faults":
    # #         ws.title = "Errors"
    # #     else:
    # #         wb.remove(ws)
    # wb.close()
    # wb2.save(newPath)

    # pass
# Spec = r"..\DataSource/CHT_SWV_GWM_P0102_2S_IMU_Test Result.xlsm"
# Spec = r"..\DataSource/CHT_SWV_GWM_P0102_2S_IMU_Test Result.xlsm"
# ExcelPath = r"E:\GitHub\MiniTool\DataSource\CHT_SWV_GWM_P0102_2S_IMU_Test Result.xlsm"
# LoopUP = "..\DataSource/P05CBID.txt"
# Df_LoopUp = ReadLoopUp(LoopUP)
# Df_PTC_Spec,LastSheet_Name,RowSize = ReadSpec_ReplaceDoorsID(Spec,Df_LoopUp)
#     ExcelAPP = win32com.client.DispatchEx('Excel.Application')
#     ExcelAPP.Visible = 0
#     ExcelAPP.DisplayAlerts = 0
#     wb = ExcelAPP.Workbooks.Open(os.path.abspath(newPath))
#     wb.Save()
#     wb.close()
#     ExcelAPP.Quit()  # 退出


