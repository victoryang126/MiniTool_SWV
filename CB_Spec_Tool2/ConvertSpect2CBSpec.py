import pandas as pd
import os
import numpy as np
# import win32com
# from pandas import ExcelWriter
# from openpyxl.utils import get_column_letter
import re
import openpyxl

def IsCBID(ID):
    """
    检查是否是codebeamer 的需求ID
    :param ID: 需求ID
    :return:True or false
    """
    # print(ID)
    pattern = re.compile(r'^\d+$')
    return pattern.match(ID)


def ConverSpecTestResult2CBCaseStatus(result):
    """

    :param result: the cell value of _VerificationStatus in TableOfContent
    :return: return the corresonding value in CB
    """
    if result.strip() == "PASSED":
        return "RESULT_PASSED"
    elif result.strip()  == 'FAILED':
        return "RESULT_FAILED"
    else:
        return "Init"

def GetCaseResult(results):
    """
    如case 数据里面有FAILED，则Failed
    如果Case 数据里面有 PASS，则Pass
    如果Case 里面有ToByAna,要如何出来
    如果其他的
    """
    # print(1)
    if "FAILED" in results:
        return "FAILED"
    elif "PASSED" in results:
        return "PASSED"
    else:
        return ""

def ConvertTsResult2CaseResult(df):
    """
    根据逻辑获取Ts Result的结果，然后准换为 Case结果

    将Comment 的数据用"," ，剔除相同元素
    """
    dict_temp = {}
    # print(df)
    for col in df:
        sdata = df[col]
        #
        # 如果title 不在sensor list 或 不等于"Time" 则保留第一行数据
        if col == "_VerificationStatus":

            # print()
            # dict_temp[col] = ""
            dict_temp[col] = GetCaseResult(sdata.values)
        elif col =="Object Text":
            dict_temp[col] = sdata.values[0]
        else:
            dict_temp[col] = ",".join([str(i) for i in sdata.drop_duplicates().values])
    return pd.Series(dict_temp)

def ReplaceCellValue(CellValue,Df_LoopUp):
    """
    根据codebeamer 的需求ID 和Doors需求ID的映射关系，替换Doors需求ID为 codebeamer 的需求ID
    :param CellValue:单元格的值
    :param Df_LoopUp:codebeamer 的需求ID 和Doors需求ID的映射DataFrame
    :return:returnValue 替换以后的值
    """
    CellValue_list = CellValue.split('\n')
    returnValue_List = []
    # reg =
    for i,cell in enumerate(CellValue_list):
        if cell.strip() in Df_LoopUp.index:
            returnValue_List.append(Df_LoopUp.loc[cell.strip(), "CBID"])
        elif cell.strip() == "":
            pass
        else:
            returnValue_List.append(cell.strip())
    # print(returnValue_List)
    # print("*"*30)
    returnValue = "\n".join(returnValue_List)
    # returnValue = returnValue.replace(" ","").strip()
    return returnValue
    # return CellValue_list

#1. 读取spec的
def ReadSpec_LastSheet_ReplaceDoorsID(Spec, Df_LoopUp):
    """
    读取specification的最后一个sheet，目前的规则，最后一个sheet应该是所有testcase 和对应需求的summary
    然后将doors需求ID 替换成 codebeamer 的需求ID，然后提供给RepaceDoorsID_SaveMacroEcel 函数 保持xlsm格式文件
    :param Spec:  specification的路径
    :param Df_LoopUp:codebeamer 的需求ID 和Doors需求ID的映射DataFrame
    :return:Df_PTC_Spec:替换需求ID以后的spec的DataFrame
            LastSheet_Name 最后一个sheet的名字
            RowSize 最后一个sheet的行大小
    """
    #获取最后一个sheet的名字
    Df_AllSheet = pd.read_excel(Spec, sheet_name=None)
    LastSheet_Name = list(Df_AllSheet)[-1]
    # print(LastSheet_Name)
    Df_spec = pd.read_excel(Spec, LastSheet_Name, dtype='str',header= 6)

    ColumnsList = ["Test Case Name","_VerifiesDOORSRequirements_SYS","_VerifiesDOORSRequirements_SW" ,"_VerifiesNonDOORSRequirements"]
    Df_spec = Df_spec[ColumnsList]
    Df_spec = Df_spec.fillna("")

    Df_spec['_VerifiesDOORSRequirements_SYS'] = Df_spec['_VerifiesDOORSRequirements_SYS'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    Df_spec['_VerifiesDOORSRequirements_SW'] = Df_spec['_VerifiesDOORSRequirements_SW'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    Df_spec['_VerifiesNonDOORSRequirements'] = Df_spec['_VerifiesNonDOORSRequirements'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)

    # print(Df_PTC_Spec)
    Df_spec.set_index("Test Case Name", inplace=True)
    #获取行数
    RowSize = len(Df_spec.index)

    # print(RowSize)
    return Df_spec,LastSheet_Name,RowSize
    #NA



# # '''
# # # 可以重新建立一个独立的进程
# def RepaceDoorsID_SaveMacroEcel(Spec,Df_PTC_Spec,LastSheet_Name,RowSize):
#
#     w = win32com.client.DispatchEx('Excel.Application')
#     w.Visible = 0
#     w.DisplayAlerts = 0
#     # print(os.path.abspath(Spec))
#     wb = w.Workbooks.Open(os.path.abspath(Spec))
#
#     try:
#         sht = wb.Worksheets(LastSheet_Name)
#         for i in range(8,RowSize + 8):
#             if sht.Cells(i, "B").Value in Df_PTC_Spec.index:
#                 # print(sht.Cells(i, "B").Value)
#                 sht.Cells(i, "M").Value = Df_PTC_Spec.loc[sht.Cells(i, "B").Value,"_VerifiesDOORSRequirements_SW"]
#     except:
#         pass
#     wb.Close(SaveChanges=1)
#     #
#     w.Quit()  # 退出

def RepaceDoorsID_SaveMacroEcel(ExcelAPP,Spec,Df_spec,LastSheet_Name,RowSize):
    """
    根据ReadSpec_LastSheet_ReplaceDoorsID 获取的数据，使用win32com 调用excel进程去保存xlsm
    格式的文件
    :param ExcelAPP:win32com.client.DispatchEx('Excel.Application')的进程
    :param Spec:
    :param Df_spec: ReadSpec_LastSheet_ReplaceDoorsID 获取最后一个sheet
    :param LastSheet_Name: ReadSpec_LastSheet_ReplaceDoorsID 获取的最后一个sheet的名字
    :param RowSize: ReadSpec_LastSheet_ReplaceDoorsID 获取的最大行数
    :return:NONE
    """
    wb = ExcelAPP.Workbooks.Open(os.path.abspath(Spec))
    try:
        sht = wb.Worksheets(LastSheet_Name)
        for i in range(8,RowSize + 8):
            if sht.Cells(i, "B").Value in Df_spec.index:
                # print(sht.Cells(i, "B").Value)
                sht.Cells(i, "L").Value = Df_spec.loc[sht.Cells(i, "B").Value, "_VerifiesDOORSRequirements_SYS"]
                sht.Cells(i, "M").Value = Df_spec.loc[sht.Cells(i, "B").Value, "_VerifiesDOORSRequirements_SW"]
                sht.Cells(i, "N").Value = Df_spec.loc[sht.Cells(i, "B").Value,"_VerifiesNonDOORSRequirements"]
    except:
        pass
    wb.Close(SaveChanges=1)

# ''

#1. 读取spec的
def ReadSpec_TableOfContent(Spec):
    """
    读取Test specification 的TableOfContent
    然后将一个 case 对应一个单元格里面需求ID
    根据单元格的需求ID的数量,
    转换成多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    :param Spec: Test specification
    :return:Df_PTC_Spec 多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    """
    print("ReadSpec_TableOfContent")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('max_colwidth', 200)
    print(1)
    Df_spec = pd.read_excel(Spec, "Table Of Contents", dtype='str')
    print(2)
    ColumnsList = ["Object Text","_VerifiesDOORSRequirements"]
    Df_spec = Df_spec[ColumnsList]
    Df_spec = Df_spec.iloc[7:,:]
    Df_spec.dropna(subset = ["Object Text"],inplace = True)

    Df_TsResult, Result_Summary = GetCaseResultFromTsResult(Spec)
    # Df_PTC_Spec = pd.concat([Df_PTC_Spec,Df_TsResult],axis=0,keys="Object Text")
    # print(Df_TsResult.head(2))
    Df_spec = pd.merge(Df_spec, Df_TsResult, on="Object Text", how="left")
    # print(Df_PTC_Spec.head(2))


    Df_spec["_VerifiesDOORSRequirements"] = Df_spec["_VerifiesDOORSRequirements"].str.strip()

    #判断 _VerifiesDOORSRequirements  split为多个元素的时候，拓展元素
    Df_ReqID = Df_spec['_VerifiesDOORSRequirements'].str.split('\n', expand=True)
    # 将单元格列表内容分开成多行， 会形成二级index
    #   0    1                          0         11
    #   11  12               =>         1         12
    Df_ReqID = Df_ReqID.stack()
    Df_ReqID = Df_ReqID.reset_index(level=1,drop=True) # 剔除二级index,
    Df_ReqID.name = "_VerifiesDOORSRequirements"
    Df_spec = Df_spec.drop(['_VerifiesDOORSRequirements'], axis=1).join(Df_ReqID) #根据index 添加 _VerifiesDOORSRequirements
    #没有数据的时候填充空白
    Df_spec.fillna("", inplace=True)

    df_Summary = pd.DataFrame({"Object Text": ["Test Environment Summary"], "Description": [Result_Summary],"_VerifiesDOORSRequirements":[" "]})
    print("*" * 30)
    print(df_Summary)
    print("*"*30)
    Df_spec = pd.concat([df_Summary, Df_spec])
    # print(Df_PTC_Spec["_VerifiesDOORSRequirements"])

    print("*"*30)
    print(Df_spec.columns)
    return Df_spec


def ReadLoopUp(LoopUP):
    """
    读取 LoopUP
    :param LoopUP:
    :return:
    """
    ColumnsList = ["CBID", "DoorsID"]
    Df_LoopUp = pd.read_csv(LoopUP, names = ColumnsList, header = None,sep = "\t",dtype= 'str')
    # if True in  Df_LoopUp.duplicated(subset=['CBID']).values:
    #     raise Exception("duplicated CBID ID")
    # elif True in  Df_LoopUp.duplicated(subset=['DoorsID']).values:
    #     raise Exception("duplicated  Doors ID")
    Df_LoopUp.set_index("DoorsID", inplace=True)
    # print(Df_LoopUp)
    # print(Df_LoopUp.loc["DES_SA2FC_SW_REQ_BOOT_585","CBID"])
    return Df_LoopUp


# # 3.生成表格
# def GenerateSpec_CB(Df_PTC_Spec,Df_LoopUp,SpecCB):
#     # ColumnsList = ["ID", "Priority","Name","Description","Pre-Action","Post-Action","Test Steps.Action","Test Steps.Expected result","Test Steps.Critical","Test Parameters","Verifies","Status","Type"]
#     ColumnsList = ["ID", "Parent", "Priority", "Name", "Description", "Pre-Action", "Post-Action",
#                    "Test Steps.Action", "Test Steps.Expected result",
#                    "Test Steps.Critical", "Test Parameters", "Verifies",
#                    "Status", "Type", "_Original_TestCaseId", "_Test_Technique",
#                    "_VerifiesNonCbRequirements", "_LastReviewDate", "_ScriptPath",
#                    "_TestType", "_RegressionStrategy", "_FunctionGroup", "_Feature",
#                    "Functional Safety Relevant", "Cyber Security Relevant"]
#     df_SpecCB = pd.DataFrame(columns= ColumnsList)
#     # print(df_SpecCB)
#     # ColumnsList_Temp = ["Name","Verifies"]
#
#     Df_PTC_Spec['_VerifiesDOORSRequirements'] = Df_PTC_Spec['_VerifiesDOORSRequirements'].apply(lambda x:x if x not in Df_LoopUp.index else "[ISSUE:" + Df_LoopUp.loc[x,"CBID"] + "]")
#
#     Df_PTC_Spec.columns = ["Name","Verifies"]
#
#     # print(Df_LoopUp.loc["undefined","CBID"])
#     df_SpecCB = df_SpecCB.append(Df_PTC_Spec)
#     #
#     # print(df_SpecCB)
#     # print(Df_PTC_Spec)
#     #CB格式要求，前面加三个空格
#     df_SpecCB["Name"] = "   " + df_SpecCB["Name"]
#     df_SpecCB.to_excel(SpecCB, sheet_name="Export", index=False)
#     # print(Df_LoopUp)

# def to_excel_auto_column_weight(df, writer, sheet_name):
#     """DataFrame保存为excel并自动设置列宽"""
#     df.to_excel(writer, sheet_name=sheet_name, index=False)
#     #  计算表头的字符宽度
#     column_widths = (
#         df.columns.to_series().apply(lambda x: len(x.encode('utf-8'))).values
#     )
#     #  计算每列的最大字符宽度
#     max_widths = (
#         df.astype(str).applymap(lambda x: len(x.encode('utf'))).agg(max).values
#     )
#     # 计算整体最大宽度
#     widths = np.max([column_widths, max_widths], axis=0)
#     # 设置列宽
#     worksheet = writer.sheets[sheet_name]
#     for i, width in enumerate(widths, 1):
#         # openpyxl引擎设置字符宽度时会缩水0.5左右个字符，所以干脆+2使左右都空出一个字宽。
#         worksheet.column_dimensions[get_column_letter(i)].width = width + 2

def GenerateSpec_CB_Init(Df_spec, SpecCB):
    """
    从Test Specification 中获取出来的DatFrame，将需求ID 写成 "[ISSUE:"  需求ID "]"的格式
    :param Df_spec:从Test Specification 中获取出来的DatFrame
    :param SpecCB: 导出的CodeBeamer的TestSpecification
    :return:df_SpecCB CodeBeamer的TestSpecification 的DataFrame
    """
    # ColumnsList = ["ID", "Priority","Name","Description","Pre-Action","Post-Action","Test Steps.Action","Test Steps.Expected result","Test Steps.Critical","Test Parameters","Verifies","Status","Type"]
    ColumnsList = ["ID", "Parent", "Priority", "Name", "Description", "Pre-Action", "Post-Action",
                   "Test Steps.Action", "Test Steps.Expected result",
                   "Test Steps.Critical", "Test Parameters", "Verifies",
                   "Status","Release", "Type", "_Original_TestCaseId", "_Test_Technique",
                   "_VerifiesNonCbRequirements", "_LastReviewDate", "_ScriptPath",
                   "_TestType", "_RegressionStrategy", "_FunctionGroup", "_Feature",
                   "Functional Safety Relevant", "Cyber Security Relevant"]
    df_SpecCB = pd.DataFrame(columns= ColumnsList)
    print(df_SpecCB)
    # Index(['Object Text', 'Description', '_VerifiesDOORSRequirements',
    #        '_VerificationStatus', 'Comments'],
    Df_spec.columns = ["Name",'Description',"Verifies","Status",'Incident ID']
    df_SpecCB["Name"] = "   " + Df_spec["Name"]
    df_SpecCB["Description"] = Df_spec["Description"]
    df_SpecCB["Status"] = Df_spec["Status"].apply(ConverSpecTestResult2CBCaseStatus)
    df_SpecCB["Incident ID"] = Df_spec["Incident ID"]
    print(Df_spec["Verifies"])
    df_SpecCB["Verifies"] = Df_spec["Verifies"].apply(lambda x: "[ISSUE:" + x + "]" if IsCBID(x) else "")
    df_SpecCB["_VerifiesNonCbRequirements"] = Df_spec["Verifies"].apply(lambda x: x if not IsCBID(x) else "")
    df_SpecCB.to_excel(SpecCB, sheet_name="Export", index=False)
    DeleteLastEmptyRow(SpecCB)
    return df_SpecCB

def DeleteLastEmptyRow(Spec):
    wb = openpyxl.load_workbook(Spec)
    ws = wb.active
    max_row = ws.max_row
    ws.delete_rows(max_row + 1)
    wb.save(Spec)




def ReadSpecCB_FromCB(SpecCB_FromCB):
    """
    解析从CodeBeamer 下载下来的TestSpec,主要是为了知道已经存在的test case 的在CodeBeamer中的Case ID
    :param SpecCB_FromCB: 从CodeBeamer 下载下来的TestSpec
    :return:Df_SpecCB    从CodeBeamer 下载下来的TestSpec 的DataFrame
            Df_ID_Case  "ID", "Parent", "Name" 形成的DataFrame
    """
    Df_SpecCB = pd.read_excel(SpecCB_FromCB, "Export", dtype='str')
    # print(Df_SpecCB.head(10))
    Df_SpecCB[["ID","Parent","Name"]] = Df_SpecCB[["ID","Parent","Name"]] .fillna(method = 'ffill')
    # print(Df_SpecCB.head(10))

    Df_ID_Case = Df_SpecCB[["ID", "Parent", "Name"]].drop_duplicates(subset=['ID'])
    Df_ID_Case.drop(0, inplace=True)
    Df_ID_Case["Name"] = Df_ID_Case["Name"].str.strip()
    Df_ID_Case.set_index("Name",inplace = True,drop = False)

    # print(Df_ID_Case.head(10))
    return Df_SpecCB,Df_ID_Case

def GenerateSpec_CB_Modify(df_SpecCB_Generate,Df_ID_Case_FromCB,Release,SpecCB_Modify):

    """
    解析从CodeBeamer 下载下来的TestSpec,
    1. 知道已经存在的test case 的在CodeBeamer中的Case ID

    :param df_SpecCB_Generate: 通过本工具生成的需要上传的Test Specification
    :param Df_ID_Case_FromCB:从CodeBeamer 下载下来的Test Specification
    :param Release 改测试case修改的版本号
    :param SpecCB_Modify: 修改后的Specification名称
    :return:NONE
    """
    print(1)
    df_SpecCB_Generate["Parent"] = Df_ID_Case_FromCB.iloc[1,1]
    #根据CodeBeamer Spec的ID给对应的Case赋值
    df_SpecCB_Generate["ID"] = df_SpecCB_Generate["Name"].apply(lambda x:Df_ID_Case_FromCB.loc[x.strip(),"ID"] if x.strip() in Df_ID_Case_FromCB.index else "")
    #对不在最新的test specification的case ，设置属性为Obsolete
    for caseName in Df_ID_Case_FromCB.index:
        if caseName not in df_SpecCB_Generate["Name"].str.strip().values:
            # print(caseName)
            Df_ID_Case_FromCB.loc[caseName,"Status"] = "Obsolete"
            Df_ID_Case_FromCB.loc[caseName,"Name"] = "    " + caseName
            df_SpecCB_Generate = df_SpecCB_Generate.append(Df_ID_Case_FromCB.loc[caseName])
            # print(Df_ID_Case_FromCB.loc[caseName])
    # print(df_SpecCB_Generate)
    print(2)
    df_SpecCB_Generate["Release"] = Release
    df_SpecCB_Generate.to_excel(SpecCB_Modify, sheet_name="Export", index=False)
    DeleteLastEmptyRow(SpecCB_Modify)

def GetInitCaseList(FinnalSpec):
    df = pd.read_excel(FinnalSpec)

    InitCaseList = df.loc[df["ID"].isnull()]["Name"].drop_duplicates().str.strip().values

    print(type(InitCaseList) )
    print(InitCaseList)
    return InitCaseList

def GetCaseResultFromTsResult(Spec):
    """
       读取Test specification 的TableOfContent
       然后将一个 case 对应一个单元格里面需求ID
       根据单元格的需求ID的数量,
       转换成多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
       :param Spec: Test specification
       :return:Df_PTC_Spec 多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
       """
    # print("ReadSpec_TableOfContent")
    # pd.set_option('display.max_columns', None)
    # pd.set_option('display.max_rows', None)
    # pd.set_option('max_colwidth', 200)

    Df_TsResult = pd.read_excel(Spec, "TS Result", dtype='str')
    # print(Df_TsResult)
    # print(Df_TsResult.head(8))
    Result_Summary = Df_TsResult.iloc[2,2]
    # print(Result_Summary)
    # df = pd.DataFrame({"Object Text": [""],"Summary":[Result_Summary]})
    # print(df)
    # Df_TsResult = pd.read_excel(Spec, "FunctionName1_Matrix", dtype='str',header=5)
    print(Result_Summary)
    Df_TsResult = Df_TsResult.iloc[6:,[3,4,6]]
    print("*"*20)
    print(Df_TsResult.head(5))
    ColumnsList = ["_VerificationStatus","Comments","Object Text" ]
    Df_TsResult.columns = ColumnsList
    Df_TsResult.dropna(subset=["Object Text"], inplace=True)
    # print(Df_TsResult["Object Text"])
    df_Group = Df_TsResult.groupby(by="Object Text", as_index=False,sort = False)
    Df_TsResult = df_Group.apply(ConvertTsResult2CaseResult)
    print("*" * 20)
    # print(Df_TsResult.head(2))
    print(Df_TsResult.head(5))
    print("#" * 20)
    return Df_TsResult,Result_Summary

if __name__ == '__main__':
    Spec = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GWM_D30_2S_IMU_Test Result.xlsm"
    Df_spec =  ReadSpec_TableOfContent(Spec)
    SpecCB =  r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GWM_D30_2S_IMU_Test Result_CB.xlsx"
    GenerateSpec_CB_Init(Df_spec, SpecCB)
    # ReadSpec_TableOfContent(Spec)
