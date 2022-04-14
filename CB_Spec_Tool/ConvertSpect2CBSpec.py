import pandas as pd
import os
import numpy as np
# import win32com
# from pandas import ExcelWriter
# from openpyxl.utils import get_column_letter
import re
import openpyxl


def ConvertCBStatus2GenerateStatus(Status,Status_CB):
    if Status.strip().upper() == "INIT":
        return Status_CB
    else:
        return Status

def ConverIsIncidentID(cellValue):
    """
    检测cell value 是否是incident ID，
    然后将内容按照[ISSUE:ID,ISSUE:ID]的格式填充
    """
    templist = ["ISSUE:"+ ID.strip() for ID in cellValue.split("\n") if re.match(r'^\d+$',ID)]
    if len(templist)>0:
        # 如果找到了元素，则按照CB的格式返回
        return "[" + ",".join(templist) + "]"
    else:
        #如果没找到，返回空
        return ""
def GetReleaseValue(CellValue):
    # print(CellValue)
    if CellValue != "":
       return CellValue.split(" - ")[1]
    else:
        return ""
def GetIncidentValue(CellValue):
    # print(CellValue)
    if CellValue != "":
       return CellValue.split(" - ")[0]
    else:
        return ""



def ConverNotIncidentID(cellValue):
    """
    检测cell value 是否是incident ID，
    然后将内容按照[ISSUE:ID,ISSUE:ID]的格式填充
    """
    templist = [ID for ID in cellValue.split("\n") if not re.match(r'^\d+$',ID)]
    return ",".join(templist)

def IsCBID(ID):
    """
    检查是否是codebeamer 的需求ID
    :param ID: 需求ID
    :return:True or false
    """
    pattern = re.compile(r'^\d+$')
    return pattern.match(ID)


def ConverSpecTestResult2CBCaseStatus(result):
    """

    :param result: the cell value of _VerificationStatus in TableOfContent
    :return: return the corresonding value in CB
    """
    if result.strip().upper() == "OK":
        return "RESULT_PASSED"
    elif result.strip().upper()  == 'NOK':
        return "RESULT_FAILED"
    else:
        return "Init"


def ReplaceCellValue(CellValue,Df_LoopUp):
    """
    根据codebeamer 的需求ID 和Doors需求ID的映射关系，替换Doors需求ID为 codebeamer 的需求ID
    :param CellValue:单元格的值
    :param Df_LoopUp:codebeamer 的需求ID 和Doors需求ID的映射DataFrame
    :return:returnValue 替换以后的值
    """
    CellValue_list = CellValue.split('\n')
    returnValue_List = []
    # reg =
    for i,cell in enumerate(CellValue_list):
        if cell.strip() in Df_LoopUp.index:
            returnValue_List.append(Df_LoopUp.loc[cell.strip(), "CBID"])
        elif cell.strip() == "":
            pass
        else:
            returnValue_List.append(cell.strip())
    # print(returnValue_List)
    # print("*"*30)
    returnValue = "\n".join(returnValue_List)
    # returnValue = returnValue.replace(" ","").strip()
    return returnValue
    # return CellValue_list

#1. 读取spec的
def ReadSpec_LastSheet_ReplaceDoorsID(Spec, Df_LoopUp):
    """
    读取specification的最后一个sheet，目前的规则，最后一个sheet应该是所有testcase 和对应需求的summary
    然后将doors需求ID 替换成 codebeamer 的需求ID，然后提供给RepaceDoorsID_SaveMacroEcel 函数 保持xlsm格式文件
    :param Spec:  specification的路径
    :param Df_LoopUp:codebeamer 的需求ID 和Doors需求ID的映射DataFrame
    :return:Df_spec:替换需求ID以后的spec的DataFrame
            LastSheet_Name 最后一个sheet的名字
            RowSize 最后一个sheet的行大小
    """
    #获取最后一个sheet的名字
    Df_AllSheet = pd.read_excel(Spec, sheet_name=None)
    LastSheet_Name = list(Df_AllSheet)[-1]
    # print(LastSheet_Name)
    Df_spec = pd.read_excel(Spec, LastSheet_Name, dtype='str',header= 6)

    ColumnsList = ["Test Case Name","_VerifiesDOORSRequirements_SYS","_VerifiesDOORSRequirements_SW" ,"_VerifiesNonDOORSRequirements"]
    Df_spec = Df_spec[ColumnsList]
    Df_spec = Df_spec.fillna("")

    Df_spec['_VerifiesDOORSRequirements_SYS'] = Df_spec['_VerifiesDOORSRequirements_SYS'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    Df_spec['_VerifiesDOORSRequirements_SW'] = Df_spec['_VerifiesDOORSRequirements_SW'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    Df_spec['_VerifiesNonDOORSRequirements'] = Df_spec['_VerifiesNonDOORSRequirements'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)

    # print(Df_spec)
    Df_spec.set_index("Test Case Name", inplace=True)
    #获取行数
    RowSize = len(Df_spec.index)

    # print(RowSize)
    return Df_spec,LastSheet_Name,RowSize
    #NA



# # '''
# # # 可以重新建立一个独立的进程
# def RepaceDoorsID_SaveMacroEcel(Spec,Df_spec,LastSheet_Name,RowSize):
#
#     w = win32com.client.DispatchEx('Excel.Application')
#     w.Visible = 0
#     w.DisplayAlerts = 0
#     # print(os.path.abspath(Spec))
#     wb = w.Workbooks.Open(os.path.abspath(Spec))
#
#     try:
#         sht = wb.Worksheets(LastSheet_Name)
#         for i in range(8,RowSize + 8):
#             if sht.Cells(i, "B").Value in Df_spec.index:
#                 # print(sht.Cells(i, "B").Value)
#                 sht.Cells(i, "M").Value = Df_spec.loc[sht.Cells(i, "B").Value,"_VerifiesDOORSRequirements_SW"]
#     except:
#         pass
#     wb.Close(SaveChanges=1)
#     #
#     w.Quit()  # 退出

def RepaceDoorsID_SaveMacroEcel(ExcelAPP,Spec,Df_spec,LastSheet_Name,RowSize):
    """
    根据ReadSpec_LastSheet_ReplaceDoorsID 获取的数据，使用win32com 调用excel进程去保存xlsm
    格式的文件
    :param ExcelAPP:win32com.client.DispatchEx('Excel.Application')的进程
    :param Spec:
    :param Df_spec: ReadSpec_LastSheet_ReplaceDoorsID 获取最后一个sheet
    :param LastSheet_Name: ReadSpec_LastSheet_ReplaceDoorsID 获取的最后一个sheet的名字
    :param RowSize: ReadSpec_LastSheet_ReplaceDoorsID 获取的最大行数
    :return:NONE
    """
    wb = ExcelAPP.Workbooks.Open(os.path.abspath(Spec))
    try:
        sht = wb.Worksheets(LastSheet_Name)
        for i in range(8,RowSize + 8):
            if sht.Cells(i, "B").Value in Df_spec.index:
                # print(sht.Cells(i, "B").Value)
                sht.Cells(i, "L").Value = Df_spec.loc[sht.Cells(i, "B").Value, "_VerifiesDOORSRequirements_SYS"]
                sht.Cells(i, "M").Value = Df_spec.loc[sht.Cells(i, "B").Value, "_VerifiesDOORSRequirements_SW"]
                sht.Cells(i, "N").Value = Df_spec.loc[sht.Cells(i, "B").Value,"_VerifiesNonDOORSRequirements"]
    except:
        pass
    wb.Close(SaveChanges=1)

# ''

#1. 读取spec的
def ReadSpec_TableOfContent(Spec):
    """
    读取Test specification 的TableOfContent
    然后将一个 case 对应一个单元格里面需求ID
    根据单元格的需求ID的数量,
    转换成多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    :param Spec: Test specification
    :return:Df_spec 多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    """
    print("ReadSpec_TableOfContent")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('max_colwidth', 200)
    print(1)
    Df_spec = pd.read_excel(Spec, "Table Of Contents", dtype='str')
    print(2)
    ColumnsList = ["Object Text","_VerificationStatus", "_VerifiesDOORSRequirements","_Comment"]
    Result_Summary = Df_spec.iloc[0,4]

    CaseTrackerID = Df_spec.iloc[2, 4]
    CB_Spec_Folder_ID = Df_spec.iloc[3, 4]
    Release = Df_spec.iloc[4,4]
    # print("*"*30)
    # print(Result_Summary)
    # print(CaseTrackerID,CB_Spec_Folder_ID,Release)
    # print("*" * 30)
    Df_spec = Df_spec[ColumnsList]
    Df_spec = Df_spec.iloc[7:,:]
    Df_spec.dropna(subset = ["Object Text"],inplace = True)
    Df_spec["_VerifiesDOORSRequirements"] = Df_spec["_VerifiesDOORSRequirements"].str.rstrip()


    #判断 _VerifiesDOORSRequirements  split为多个元素的时候，拓展元素
    Df_ReqID = Df_spec['_VerifiesDOORSRequirements'].str.split('\n', expand=True)
    # 将单元格列表内容分开成多行， 会形成二级index
    #   0    1                          0         11
    #   11  12               =>         1         12
    Df_ReqID = Df_ReqID.stack()
    Df_ReqID = Df_ReqID.reset_index(level=1,drop=True) # 剔除二级index,
    Df_ReqID.name = "_VerifiesDOORSRequirements"
    Df_specTemp1 = Df_spec.drop(['_VerifiesDOORSRequirements', "_Comment"], axis=1).join(Df_ReqID)  # 根据index 添加 _VerifiesDOORSRequirements
    Df_specTemp1['_Comment'] = ""

    #将Incident ID分行,然后转换为DataFrame,_VerifiesDOORSRequirements 列设置为空
    Df_IssueID = Df_spec['_Comment'].str.split('\n', expand=True)
    Df_IssueID = Df_IssueID.stack()
    Df_IssueID = Df_IssueID.reset_index(level=1,drop=True) # 剔除二级index,
    Df_IssueID.name = "_Comment"
    data = {"_Comment" :Df_IssueID.values}
    # print(data)
    Df_IssueID = pd.DataFrame(data =data ,index= Df_IssueID.index)
    Df_IssueID["_VerifiesDOORSRequirements"] = ""
    Df_specTemp2 = Df_spec.drop(['_Comment','_VerifiesDOORSRequirements'], axis=1).join(Df_IssueID,how="inner")  # 根据index 添加 _VerifiesDOORSRequirements

    #合并表格，然后按照index 重新排序
    Df_spec = pd.concat([Df_specTemp1,Df_specTemp2],axis=0)
    # 上面的步骤会把Df_IssueID  放到最后，需要通过sort index将 对应case的ID单元啦上去
    Df_spec.fillna("", inplace=True)
    Df_spec.sort_index(inplace=True) #

    #_VerifiesDOORSRequirements 是换行的过滤
    # Df_spec = Df_spec[Df_spec['_VerifiesDOORSRequirements'].str.strip() != ""]
    # Df_spec = Df_spec[Df_spec['_Comment'].str.strip() != ""]



    #根据版本好添加summary
    df_Summary = pd.DataFrame({"Object Text": ["Test Summary in " + Release], "Description": [Result_Summary],"_VerifiesDOORSRequirements":[" "],"Type":["Information"]})
    print("*" * 30)
    # print(Df_spec.head(20))
    print("*"*30)
    Df_spec = pd.concat([df_Summary, Df_spec])
    Df_spec.fillna("",inplace=True)
    print(Df_spec.columns)
    # Index(['Object Text', 'Description', '_VerifiesDOORSRequirements', 'Type',
    #        '_VerificationStatus', '_Comment'],
    # //Index(['Object Text', 'Description', '_VerifiesDOORSRequirements',
    #    '_VerificationStatus', '_Comment'],
    #   dtype='object')
    return Df_spec,CaseTrackerID,CB_Spec_Folder_ID,Release


def ReadLoopUp(LoopUP):
    """
    读取 LoopUP
    :param LoopUP:
    :return:
    """
    ColumnsList = ["CBID", "DoorsID"]
    Df_LoopUp = pd.read_csv(LoopUP, names = ColumnsList, header = None,sep = "\t",dtype= 'str')
    # if True in  Df_LoopUp.duplicated(subset=['CBID']).values:
    #     raise Exception("duplicated CBID ID")
    # elif True in  Df_LoopUp.duplicated(subset=['DoorsID']).values:
    #     raise Exception("duplicated  Doors ID")
    Df_LoopUp.set_index("DoorsID", inplace=True)
    # print(Df_LoopUp)
    # print(Df_LoopUp.loc["DES_SA2FC_SW_REQ_BOOT_585","CBID"])
    return Df_LoopUp


# # 3.生成表格
# def GenerateSpec_CB(Df_spec,Df_LoopUp,SpecCB):
#     # ColumnsList = ["ID", "Priority","Name","Description","Pre-Action","Post-Action","Test Steps.Action","Test Steps.Expected result","Test Steps.Critical","Test Parameters","Verifies","Status","Type"]
#     ColumnsList = ["ID", "Parent", "Priority", "Name", "Description", "Pre-Action", "Post-Action",
#                    "Test Steps.Action", "Test Steps.Expected result",
#                    "Test Steps.Critical", "Test Parameters", "Verifies",
#                    "Status", "Type", "_Original_TestCaseId", "_Test_Technique",
#                    "_VerifiesNonCbRequirements", "_LastReviewDate", "_ScriptPath",
#                    "_TestType", "_RegressionStrategy", "_FunctionGroup", "_Feature",
#                    "Functional Safety Relevant", "Cyber Security Relevant"]
#     df_SpecCB = pd.DataFrame(columns= ColumnsList)
#     # print(df_SpecCB)
#     # ColumnsList_Temp = ["Name","Verifies"]
#
#     Df_spec['_VerifiesDOORSRequirements'] = Df_spec['_VerifiesDOORSRequirements'].apply(lambda x:x if x not in Df_LoopUp.index else "[ISSUE:" + Df_LoopUp.loc[x,"CBID"] + "]")
#
#     Df_spec.columns = ["Name","Verifies"]
#
#     # print(Df_LoopUp.loc["undefined","CBID"])
#     df_SpecCB = df_SpecCB.append(Df_spec)
#     #
#     # print(df_SpecCB)
#     # print(Df_spec)
#     #CB格式要求，前面加三个空格
#     df_SpecCB["Name"] = "   " + df_SpecCB["Name"]
#     df_SpecCB.to_excel(SpecCB, sheet_name="Export", index=False)
#     # print(Df_LoopUp)

# def to_excel_auto_column_weight(df, writer, sheet_name):
#     """DataFrame保存为excel并自动设置列宽"""
#     df.to_excel(writer, sheet_name=sheet_name, index=False)
#     #  计算表头的字符宽度
#     column_widths = (
#         df.columns.to_series().apply(lambda x: len(x.encode('utf-8'))).values
#     )
#     #  计算每列的最大字符宽度
#     max_widths = (
#         df.astype(str).applymap(lambda x: len(x.encode('utf'))).agg(max).values
#     )
#     # 计算整体最大宽度
#     widths = np.max([column_widths, max_widths], axis=0)
#     # 设置列宽
#     worksheet = writer.sheets[sheet_name]
#     for i, width in enumerate(widths, 1):
#         # openpyxl引擎设置字符宽度时会缩水0.5左右个字符，所以干脆+2使左右都空出一个字宽。
#         worksheet.column_dimensions[get_column_letter(i)].width = width + 2

def GenerateSpec_CB_Init(Df_spec,Release, SpecCB):
    """
    从Test Specification 中获取出来的DatFrame，将需求ID 写成 "[ISSUE:"  需求ID "]"的格式
    :param Df_spec:从Test Specification 中获取出来的DatFrame
    :param SpecCB: 导出的CodeBeamer的TestSpecification
    :return:df_SpecCB CodeBeamer的TestSpecification 的DataFrame
    """
    # ColumnsList = ["ID", "Priority","Name","Description","Pre-Action","Post-Action","Test Steps.Action","Test Steps.Expected result","Test Steps.Critical","Test Parameters","Verifies","Status","Type"]
    ColumnsList = ["ID", "Parent", "Priority", "Name", "Description", "Pre-Action", "Post-Action",
                   "Test Steps.Action", "Test Steps.Expected result",
                   "Test Steps.Critical", "Test Parameters", "Verifies",
                   "Status","Release", "Type", "_Original_TestCaseId", "_Test_Technique",
                   "_VerifiesNonCbRequirements", "_LastReviewDate", "_ScriptPath",
                   "_TestType", "_RegressionStrategy", "_FunctionGroup", "_Feature",
                   "Functional Safety Relevant", "Cyber Security Relevant","Reserve_1"]
    df_SpecCB = pd.DataFrame(columns= ColumnsList)

    print("GenerateSpec_CB_Init")
    # Index(['Object Text', 'Description', '_VerifiesDOORSRequirements', 'Type',
    #        '_VerificationStatus', '_Comment'],
    # //Index(['Object Text', 'Description', '_VerifiesDOORSRequirements',
    #    '_VerificationStatus', '_Comment'],
    #   dtype='object')
    Df_spec.columns =  ["Name", 'Description', "Verifies",'Type',"Status", 'Incident ID']

    df_SpecCB["Name"] = "    " + Df_spec["Name"]
    df_SpecCB["Description"] = Df_spec["Description"]
    df_SpecCB["Status"] = Df_spec["Status"].apply(ConverSpecTestResult2CBCaseStatus)
    df_SpecCB["Verifies"] = Df_spec["Verifies"].apply(lambda x: "[ISSUE:" + x + "]" if IsCBID(x) else "")

    df_SpecCB["Type"] = Df_spec["Type"]

    df_SpecCB["_VerifiesNonCbRequirements"] = Df_spec["Verifies"].apply(lambda x: x if not IsCBID(x) else "")
    df_SpecCB["Incident ID"] = Df_spec["Incident ID"].apply(lambda x: "[ISSUE:" + x + "]" if IsCBID(x) else "")
    df_SpecCB["Reserve_1"] = Df_spec["Incident ID"].apply(lambda x: x if not IsCBID(x) else "")
    # df_SpecCB["Release"] = Release
    # print(df_SpecCB[["Name","Verifies"]])
    df_SpecCB.to_excel(SpecCB, sheet_name="Export", index=False)
    DeleteLastEmptyRow(SpecCB)
    return df_SpecCB

def DeleteLastEmptyRow(Spec):
    wb = openpyxl.load_workbook(Spec)
    ws = wb.active
    max_row = ws.max_row
    ws.delete_rows(max_row + 1)
    wb.save(Spec)




def ReadSpecCB_FromCB(SpecCB_FromCB):
    """
    解析从CodeBeamer 下载下来的TestSpec,主要是为了知道已经存在的test case 的在CodeBeamer中的Case ID
    :param SpecCB_FromCB: 从CodeBeamer 下载下来的TestSpec
    :return:Df_SpecCB    从CodeBeamer 下载下来的TestSpec 的DataFrame
            Df_ID_Case  "ID", "Parent", "Name" 形成的DataFrame
    """
    print("&"*30 + "ReadSpecCB_FromCB" + "&"*30)
    Df_SpecCB = pd.read_excel(SpecCB_FromCB, "Export", dtype='str')
    # print(Df_SpecCB.head(10))
    Df_SpecCB[["ID","Parent","Name"]] = Df_SpecCB[["ID","Parent","Name"]] .fillna(method = 'ffill')
    # print(Df_SpecCB.head(10))

    Df_ID_Case = Df_SpecCB[["ID", "Parent", "Name","Type","Status"]].drop_duplicates(subset=['ID'])
    Df_ID_Case.drop(0, inplace=True)
    Df_ID_Case["Name"] = Df_ID_Case["Name"].str.strip()
    Df_ID_Case.set_index("Name",inplace = True,drop = False)


    # print(Df_ID_Case.columns)
    return Df_SpecCB,Df_ID_Case

def ReadSpecCB_FromCB2(SpecCB_FromCB):
    """
    解析从CodeBeamer 下载下来的TestSpec,主要是为了知道已经存在的test case 的在CodeBeamer中的Case ID
    :param SpecCB_FromCB: 从CodeBeamer 下载下来的TestSpec
    :return:Df_SpecCB    从CodeBeamer 下载下来的TestSpec 的DataFrame
            Df_ID_Case  "ID", "Parent", "Name" 形成的DataFrame
    """
    print("&"*30 + "ReadSpecCB_FromCB2" + "&"*30)
    Df_SpecCB = pd.read_excel(SpecCB_FromCB, "Export",dtype=object)
    # print(Df_SpecCB["ID"])
    # print(Df_SpecCB.head(10))
    Df_SpecCB[["ID","Parent","Name"]] = Df_SpecCB[["ID","Parent","Name"]] .fillna(method = 'ffill')
    # print(Df_SpecCB.head(10))
    #获取版本号和issue ID
    Df_SpecCB= Df_SpecCB.fillna({"Release":"","Incident ID":"","Status":"","Verifies":""})
    Df_SpecCB["Release"] = Df_SpecCB["Release"].apply(GetReleaseValue)
    Df_SpecCB["Incident ID"] = Df_SpecCB["Incident ID"].apply(GetIncidentValue)
    Df_SpecCB["Verifies"] = Df_SpecCB["Verifies"].apply(GetIncidentValue)
    Df_SpecCB =  Df_SpecCB[["ID","Parent","Name","Verifies","Status","Type","Release","Incident ID"]]
    # Df_SpecCB.drop(0, inplace=True)


    Df_ID_Case = Df_SpecCB[["ID", "Parent", "Name","Type","Status"]].drop_duplicates(subset=['ID'])

    Df_ID_Case.drop(0, inplace=True)
    Df_ID_Case["Name"] = Df_ID_Case["Name"].str.strip()
    Df_ID_Case.set_index("Name",inplace = True,drop = False)
    Df_ID_Case = Df_ID_Case.astype({"ID":"int64","Parent":"int64"})
    # print( Df_ID_Case.astype({"ID":"int64","Parent":"int64"})["ID"].dtypes)
    # Df_ID_Case.astype("ID":)
    # print(Df_ID_Case["ID"])
    # Df_ID_Case["ID"] = Df_ID_Case["ID"].str.replace(".0","")
    # Df_ID_Case["Parent"] = Df_ID_Case["Parent"].str.replace(".0", "")
    # Df_SpecCB = Df_SpecCB[(Df_SpecCB["Incident ID"] !="") | (Df_SpecCB["Release"] !="")]
    Df_SpecCB = Df_SpecCB[Df_SpecCB["Status"] != "Obsolete"] #删除status 部位obsolete的case
    # print(Df_ID_Case)
    # print(Df_ID_Case.columns)

    Df_SpecCB.drop(0, inplace=True)
    # print(Df_SpecCB.iloc[0, 2])
    return Df_SpecCB,Df_ID_Case

def GenerateSpec_CB_Modify(df_SpecCB_Generate,Df_ID_Case_FromCB,Release,SpecCB_Modify):

    """
    解析从CodeBeamer 下载下来的TestSpec,
    1. 知道已经存在的test case 的在CodeBeamer中的Case ID

    :param df_SpecCB_Generate: 通过本工具生成的需要上传的Test Specification
    :param Df_ID_Case_FromCB:从CodeBeamer 下载下来的Test Specification
    :param Release 该测试case修改的版本号，CB版本号
    :param SpecCB_Modify: 修改后的Specification名称
    :return:NONE
    """
    print("GenerateSpec_CB_Modify")
    df_SpecCB_Generate["Parent"] = Df_ID_Case_FromCB.iloc[1,1]
    #根据CodeBeamer Spec的ID给对应的Case赋值
    df_SpecCB_Generate["ID"] = df_SpecCB_Generate["Name"].apply(lambda x:Df_ID_Case_FromCB.loc[x.strip(),"ID"] if x.strip() in Df_ID_Case_FromCB.index else "")
    #对不在最新的test specification的case ，设置属性为Obsolete
    for caseName in Df_ID_Case_FromCB.index:

        # 如果 属性为information,则跳过，继续判断
        if Df_ID_Case_FromCB.loc[caseName,"Type"] == "Information":
            # print("2")
            continue;
        if caseName not in df_SpecCB_Generate["Name"].str.strip().values:
            # print(caseName)
            # Df_ID_Case_FromCB.loc[caseName,"Status"] = "Obsolete"
            Df_ID_Case_FromCB.loc[caseName,"Name"] = "    " + caseName
            df_SpecCB_Generate = df_SpecCB_Generate.append(Df_ID_Case_FromCB.loc[caseName])
            # print(Df_ID_Case_FromCB.loc[caseName])
    # print(df_SpecCB_Generate)
    print(2)
    df_SpecCB_Generate["Release"] = Release
    df_SpecCB_Generate.to_excel(SpecCB_Modify, sheet_name="Export", index=False)
    DeleteLastEmptyRow(SpecCB_Modify)



def GenerateSpec_CB_Modify2(df_SpecCB_Generate,Df_ID_Case_FromCB,Df_SpecCB_FromCB,Release,SpecCB_Modify):

    """
    解析从CodeBeamer 下载下来的TestSpec,
    1. 知道已经存在的test case 的在CodeBeamer中的Case ID

    :param df_SpecCB_Generate: 通过本工具生成的需要上传的Test Specification
    :param Df_ID_Case_FromCB:从CodeBeamer 下载下来的Test Specification
    :param Release 该测试case修改的版本号，CB版本号
    :param SpecCB_Modify: 修改后的Specification名称
    :return:NONE
    """
    print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&GenerateSpec_CB_Modify2&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    df_SpecCB_Generate["Parent"] = Df_ID_Case_FromCB.iloc[1,1]
    #根据CodeBeamer Spec的ID给对应的Case赋值
    # print(df_SpecCB_Generate.index)
    df_SpecCB_Generate["ID"] = df_SpecCB_Generate["Name"].apply(lambda x:Df_ID_Case_FromCB.loc[x.strip(),"ID"] if x.strip() in Df_ID_Case_FromCB.index else "")
    # print(df_SpecCB_Generate[['ID', "Name"]])
    # print("*" * 40)
    # print(Df_ID_Case_FromCB[['ID', "Name"]])
    #对不在最新的test specification的case ，设置属性为Obsolete
    for caseName in Df_ID_Case_FromCB.index:

        # 如果 属性为information,则跳过，继续判断
        if Df_ID_Case_FromCB.loc[caseName,"Type"] == "Information":
            # print("2")
            continue;
        if caseName not in df_SpecCB_Generate["Name"].str.strip().values:
            # print(caseName)
            Df_ID_Case_FromCB.loc[caseName,"Status"] = "Obsolete"
            Df_ID_Case_FromCB.loc[caseName,"Name"] = "    " + caseName
            df_SpecCB_Generate = df_SpecCB_Generate.append(Df_ID_Case_FromCB.loc[caseName])

            #从Df_SpecCB_FromCB 中删除掉这个case
            Df_SpecCB_FromCB = Df_SpecCB_FromCB[Df_SpecCB_FromCB["Name"].str.strip() != caseName]
            # print(Df_SpecCB_FromCB["Name"])
            print("*" * 40)
            # print(Df_SpecCB_FromCB["Name","Status"])
            # print(Df_ID_Case_FromCB.loc[caseName])
    df_SpecCB_Generate["Release"] = Release
    # 、**********************************判断DF_Generate 里面的case ID是否有新增加**********************************
    # Verify 可能会有空的数据，导致Df_Verify_Generate_2 里面没数据,所以需要剔除这个部分的数据
    Df_Verify_CB = Df_SpecCB_FromCB[["Name", "Verifies"]]
    Df_Verify_CB.dropna(subset=["Name", "Verifies"], inplace=True)
    Df_Verify_CB["Name"] = Df_Verify_CB["Name"].str.strip()
    Df_Verify_CB_2 = Df_Verify_CB["Name"].str.strip() + "," + Df_Verify_CB["Verifies"]

    Df_Verify_Generate = df_SpecCB_Generate[["Name", "Verifies"]]
    Df_Verify_Generate.dropna(subset=["Name", "Verifies"], inplace=True)
    Df_Verify_Generate["Name"] = Df_Verify_Generate["Name"].str.strip()
    Df_Verify_Generate_2 = Df_Verify_Generate["Name"] + "," + Df_Verify_Generate["Verifies"]

    # print(Df_Verify_Generate_2.isin(Df_Verify_CB_2))
    Df_New_Verify = Df_Verify_Generate[~Df_Verify_Generate_2.isin(Df_Verify_CB_2)]
    # print(Df_New_Verify)

    # *****************************获取之前的状态
    #先获取CB的状态，确保case名字唯一
    Df_Status_FromCB = Df_SpecCB_FromCB[["Name","Status"]]
    Df_Status_FromCB["Name"] = Df_Status_FromCB["Name"].str.strip()
    Df_Status_FromCB.drop_duplicates(keep="first",inplace = True,subset=['Name'])
    #然后获取Generate的状态
    Df_Status_Generate = df_SpecCB_Generate[["Name","Status","Release"]]
    Df_Status_Generate["Name"] = Df_Status_Generate["Name"].str.strip()
    Df_Status_temp  = Df_Status_Generate.merge(Df_Status_FromCB,left_on="Name",right_on="Name",suffixes=["","_CB"],how="left")
    Df_Status_temp["Status_CB"] = Df_Status_temp["Status_CB"].fillna("Init")
    # Df_Status_temp.set_index("Name",inplace=True)
    # for i in
    # Df_Status_temp["Status"] = Df_Status_temp["Status"].apply(ConvertCBStatus2GenerateStatus,Status_CB=Df_Status_temp["Status_CB"])
    # Df_Status_temp["Status"] = Df_Status_temp["Status"].apply(lambda x: x if x.strip().upper =="INIT"  else Df_ID_Case_FromCB.loc[x, "Status"])

    for i in Df_Status_temp.index:
        # print(Df_Status_temp.loc[i,"Status"].strip().upper == "INIT")
        #只要是init就把case的 Release给干掉
        if Df_Status_temp.loc[i,"Status"].strip().upper() == "INIT":
            Df_Status_temp.loc[i, "Release"] = ""
            # print(Df_Status_temp.loc[i,"Name"])
            # 如果这个case 没有新增加的需求ID，则使用之前的case状态
            # print(Df_Status_temp.loc[i,"Name"].strip() not in Df_New_Verify["Name"].values)
            if Df_Status_temp.loc[i,"Name"].strip() not in Df_New_Verify["Name"].values:
                Df_Status_temp.loc[i, "Status"]=  Df_Status_temp.loc[i,"Status_CB"]





    print(2)
    # print(df_SpecCB_Generate[[ "Status"]])
    print("*"*40)
    df_SpecCB_Generate.index = Df_Status_temp.index
    df_SpecCB_Generate["Release"] = Df_Status_temp["Release"]
    df_SpecCB_Generate["Status"] = Df_Status_temp["Status"]
    # print(df_SpecCB_Generate[["Status"]] )


    Df_SpecCB_FromCB["Name"] = "    " + Df_SpecCB_FromCB["Name"]
    Df_SpecCB_FromCB["Status"] = ""
    Df_SpecCB_Temp = Df_SpecCB_FromCB[(Df_SpecCB_FromCB["Incident ID"] != "") | (Df_SpecCB_FromCB["Release"] != "")]
    df_SpecCB_Generate = pd.concat([df_SpecCB_Generate,Df_SpecCB_Temp])

    # Name 按照升序，Status按照降序，Status 为空的必须放在后面，这个部分的行主要是为了添加之前CB 上的Release号和Issue
    df_SpecCB_Generate.sort_values(by=["Name","Status"],ascending=[True,False],inplace = True)
    # print(df_SpecCB_Generate[['ID', "Name"]])
    df_SpecCB_Generate.to_excel(SpecCB_Modify, sheet_name="Export", index=False)
    DeleteLastEmptyRow(SpecCB_Modify)

def GetInitCaseList(FinnalSpec):
    df = pd.read_excel(FinnalSpec)

    InitCaseList = df.loc[df["ID"].isnull()]["Name"].drop_duplicates().str.strip().values

    print(type(InitCaseList) )
    print(InitCaseList)
    return InitCaseList


if __name__ == '__main__':
    pass
    Spec = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GMW_D30_2S_DCS_Test_Result.xlsm"
    Df_spec,CaseTrackerID,CB_Spec_Folder_ID,Release= ReadSpec_TableOfContent(Spec)
    SpecCB = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GMW_D30_2S_DCS_Test_Result_CB.xlsx"
    SpecCB_FromCB = r"C:\Users\victor.yang\Desktop\Work\CB\84194_GWM_D30_RCS_SC2_2S - TC_L30_DES_SW_Test_Cases (4).xlsx"
    CB_Spec_Generate = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GMW_D30_2S_DCS_Test_Result_CodeBeamer.xlsx"
    GenerateSpec_CB_Init(Df_spec,Release,CB_Spec_Generate)
    df_SpecCB_FromCB, Df_ID_Case_FromCB =  ReadSpecCB_FromCB2(SpecCB_FromCB)
    df_SpecCB_Generate = pd.read_excel(CB_Spec_Generate, "Export")
    SpecCB_Modify = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GMW_D30_2S_DCS_Test_Result_CodeBeamer_Modify.xlsx"


    GenerateSpec_CB_Modify2(df_SpecCB_Generate, Df_ID_Case_FromCB, df_SpecCB_FromCB,Release,
                                    SpecCB_Modify)
    # GenerateSpec_CB_Init(Df_spec,Release,SpecCB)
    # ReadSpec_TableOfContent(Spec)
    # pass
    # Spec = "C:\Users\victor.yang\Desktop\Work\CB"
    # LoopUP = "..\DataSource/P05CBID.txt"
    # SpecCB_Generate = "..\Data/CHT_SWV_GWM_P0102_2S_IMU_Test Result_CodeBeamer.xlsx"
    # SpecCB_Modify = "..\Data/CHT_SWV_GWM_P0102_2S_IMU_Test Result_CodeBeamer.xlsx"
    # SpecCB_FromCB = "..\Data/84177 _GWM_P05_IMU_Test case.xlsx"
    # Df_LoopUp = ReadLoopUp(LoopUP)
    #
    # Df_spec = ReadSpec_TableOfContent(Spec)
    #
    # df_SpecCB_Generate = GenerateSpec_CB_Init(Df_spec, SpecCB_Generate)
    # df_SpecCB_FromCB,Df_ID_Case_FromCB = ReadSpecCB_FromCB(SpecCB_FromCB)
    # GenerateSpec_CB_Modify(df_SpecCB_Generate, Df_ID_Case_FromCB, SpecCB_Modify)
    # a = r"C:/Users/victor.yang/Downloads\CHT_System_Validation_GWM_D30_CANC_Test Specification_CodeBeamer.xlsx"
    # CodeBeamer_Spec = "E:\Project_Test\Geely_Geea2_HX11\DCS\CHT_System_Validation_Chery_T26_CANC_Test Specification_CodeBeamer.xlsx"
    # GetInitCaseList(CodeBeamer_Spec)
    # df = pd.read_excel(a)
    # print(df.index)
    # print(df["Verifies"])

    # APISheet.
    # print(row)