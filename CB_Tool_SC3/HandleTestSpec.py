import pandas as pd
import os
import numpy as np
from openpyxl.styles import Font
from CB_Tool_SC3.CodeBeamer import CodeBeamer
# import win32com
# from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re
import openpyxl

def GetCBID(x,df):
    # lambda x: Df_ID_Case_FromCB.loc[x.strip(), "ID"] if x.strip() in Df_ID_Case_FromCB.index else ""
    # print(x)
    if x.strip() in df.index:
        # print(df.loc[x.strip(), "ID"])
        return df.loc[x.strip(), "ID"]
    else:
        print("NONE")
        return ""


def ConvertCBStatus2GenerateStatus(Status,Status_CB):
    if Status.strip().upper() == "INIT":
        return Status_CB
    else:
        return Status

def ConverIsIncidentID(cellValue):
    """
    检测cell value 是否是incident ID，
    然后将内容按照[ISSUE:ID,ISSUE:ID]的格式填充
    """
    templist = ["ISSUE:"+ ID.strip() for ID in cellValue.split("\n") if re.match(r'^\d+$',ID)]
    if len(templist)>0:
        # 如果找到了元素，则按照CB的格式返回
        return "[" + ",".join(templist) + "]"
    else:
        #如果没找到，返回空
        return ""
def GetReleaseValue(CellValue):
    # print(cellvalue)
    if CellValue != "":
       return CellValue.split(" - ")[1]
    else:
        return ""
def GetIncidentValue(CellValue):
    # print(cellvalue)
    if CellValue != "":
       return CellValue.split(" - ")[0]
    else:
        return ""



def ConverNotIncidentID(cellValue):
    """
    检测cell value 是否是incident ID，
    然后将内容按照[ISSUE:ID,ISSUE:ID]的格式填充
    """
    templist = [ID for ID in cellValue.split("\n") if not re.match(r'^\d+$',ID)]
    return ",".join(templist)

def IsCBID(ID):
    """
    检查是否是codebeamer 的需求ID
    :param ID: 需求ID
    :return:True or false
    """
    pattern = re.compile(r'^\d+$')
    return pattern.match(ID)


def ConverSpecTestResult2CBCaseStatus(result):
    """

    :param result: the cell value of _VerificationStatus in TableOfContent
    :return: return the corresonding value in CB
    """
    if result.strip().upper() == "OK":
        return "RESULT_PASSED"
    elif result.strip().upper()  == 'NOK':
        return "RESULT_FAILED"
    else:
        return "Init"


def ReplaceCellValue(CellValue,Df_LoopUp):
    """
    根据codebeamer 的需求ID 和Doors需求ID的映射关系，替换Doors需求ID为 codebeamer 的需求ID
    :param CellValue:单元格的值
    :param Df_LoopUp:codebeamer 的需求ID 和Doors需求ID的映射DataFrame
    :return:returnValue 替换以后的值
    """
    CellValue_list = CellValue.split('\n')
    returnValue_List = []
    # reg =
    for i,cell in enumerate(CellValue_list):
        if cell.strip() in Df_LoopUp.index:
            returnValue_List.append(Df_LoopUp.loc[cell.strip(), "CBID"])
        elif cell.strip() == "":
            pass
        else:
            returnValue_List.append(cell.strip())
    # print(returnValue_List)
    # print("*"*30)
    returnValue = "\n".join(returnValue_List)
    # returnValue = returnValue.replace(" ","").strip()
    return returnValue
    # return CellValue_list

#1. 读取spec的
def ReadSpec_LastSheet_ReplaceDoorsID(Spec, Df_LoopUp):
    """
    读取specification的最后一个sheet，目前的规则，最后一个sheet应该是所有testcase 和对应需求的summary
    然后将doors需求ID 替换成 codebeamer 的需求ID，然后提供给RepaceDoorsID_SaveMacroEcel 函数 保持xlsm格式文件
    :param Spec:  specification的路径
    :param Df_LoopUp:codebeamer 的需求ID 和Doors需求ID的映射DataFrame
    :return:Df_PTC_Spec:替换需求ID以后的spec的DataFrame
            LastSheet_Name 最后一个sheet的名字
            RowSize 最后一个sheet的行大小
    """
    #获取最后一个sheet的名字
    Df_AllSheet = pd.read_excel(Spec, sheet_name=None)
    LastSheet_Name = list(Df_AllSheet)[-1]
    # print(LastSheet_Name)
    Df_spec = pd.read_excel(Spec, LastSheet_Name, dtype='str',header= 6)

    ColumnsList = ["Test Case Name","_VerifiesDOORSRequirements_SYS","_VerifiesDOORSRequirements_SW" ,"_VerifiesNonDOORSRequirements"]
    Df_spec = Df_spec[ColumnsList]
    Df_spec = Df_spec.fillna("")

    Df_spec['_VerifiesDOORSRequirements_SYS'] = Df_spec['_VerifiesDOORSRequirements_SYS'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    Df_spec['_VerifiesDOORSRequirements_SW'] = Df_spec['_VerifiesDOORSRequirements_SW'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)
    Df_spec['_VerifiesNonDOORSRequirements'] = Df_spec['_VerifiesNonDOORSRequirements'].apply(ReplaceCellValue,Df_LoopUp = Df_LoopUp)

    # print(Df_PTC_Spec)
    Df_spec.set_index("Test Case Name", inplace=True)
    #获取行数
    RowSize = len(Df_spec.index)

    # print(RowSize)
    return Df_spec,LastSheet_Name,RowSize
    #NA



def ConvertVerifisPTCFormat_ToCBFormat(CellValue,CBReqFlag):
    # print(type(cellvalue))

    if CBReqFlag:
        if CellValue != "":
            print(1)
            CellValue = CellValue.strip()
            cell_vaule_list = CellValue.split("\n")
            cell_vaule_list = [ "[ISSUE:" + x + "]" for x in cell_vaule_list if IsCBID(x)]
            CellValue= "\n".join(cell_vaule_list)
            return CellValue
        else:
            return CellValue
    else: # 获取非CBID的需求
        if CellValue != "":
            print(1)
            CellValue = CellValue.strip()
            cell_vaule_list = CellValue.split("\n")
            cell_vaule_list = [x for x in cell_vaule_list if not IsCBID(x)]
            CellValue = ",".join(cell_vaule_list)
            return CellValue
        else:
            return CellValue



def ConvertIncidentPTCFormat_ToCBFormat(CellValue):
    #不为空,则处理，否则不处理
    if CellValue != "":
        CellValue = CellValue.strip()
        cell_vaule_list = CellValue.split("\n")
        if IsCBID(cell_vaule_list[0]): #如果第一个元素不为CBID,则也是comment,返回原来的值
            cell_vaule_list = ["[ISSUE:" + x + "]" for x in cell_vaule_list if IsCBID(x)]
            CellValue = "\n".join(cell_vaule_list)
        return CellValue
    else:
        return CellValue


#1. 读取spec的
def ReadSpec_TableOfContent_SC3(Spec,CodeBeamer_Obj):
    """
    读取Test specification 的TableOfContent
    然后将一个 case 对应一个单元格里面需求ID
    根据单元格的需求ID的数量,
    转换成多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    :param Spec: Test specification
    :return:Df_PTC_Spec 多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    """
    print("ReadSpec_TableOfContent_SC3")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('max_colwidth', 200)
    print("ReadSpec_TableOfContent_SC3 Spec:" + Spec)
    df_ptc_spec = pd.read_excel(Spec, "Table Of Contents", dtype='str')

    cols = df_ptc_spec.columns

    if "Test Method" not in cols:
        df_ptc_spec["Test Method"] = ""

    ColumnsList = ["Object Text", "_VerificationStatus", "_VerifiesDOORSRequirements", "Test Method"]

    CodeBeamer_Obj.Result_Summary= df_ptc_spec.iloc[0, 4]
    CodeBeamer_Obj.TestRun_TrackerName = df_ptc_spec.iloc[1, 4]
    CodeBeamer_Obj.CaseTrackerID = df_ptc_spec.iloc[2, 4]
    CodeBeamer_Obj.CaseFolderID = df_ptc_spec.iloc[3, 4]
    CodeBeamer_Obj.Release = df_ptc_spec.iloc[4,4]
    # print("*"*30)
    # print(Result_Summary)
    # print(CaseTrackerID,CB_Spec_Folder_ID,Release)
    # print("*" * 30)
    df_ptc_spec = df_ptc_spec[ColumnsList]
    df_ptc_spec = df_ptc_spec.iloc[7:,:]
    df_ptc_spec.dropna(subset = ["Object Text"],inplace = True)
    df_ptc_spec["_VerifiesDOORSRequirements"] = df_ptc_spec["_VerifiesDOORSRequirements"].str.rstrip()



    df_ptc_spec.columns = ["Name","Status","Verifies","Test Method"]
    df_ptc_spec.fillna("",inplace = True)
    df_ID = df_ptc_spec["Verifies"]
    df_ptc_spec["Verifies"] = df_ID.apply(ConvertVerifisPTCFormat_ToCBFormat,CBReqFlag = True)
    df_ptc_spec["_VerifiesNonCbRequirements"] = df_ID.apply(ConvertVerifisPTCFormat_ToCBFormat,CBReqFlag =False)
    df_ptc_spec["Test Method"] = df_ptc_spec["Test Method"].apply(lambda x:x.replace("\n",","))

    print(df_ptc_spec)
    # df_spec.to_excel(r"C:\Users\victor.yang\Desktop\Work\CB\SpecTemplate\Test.xlsx", sheet_name="Export", index=False)
    return df_ptc_spec





def GenerateSpec_CB_Init(Df_PTC_Spec, SpecCB,CodeBeamer_Obj):
    """
    从Test Specification 中获取出来的DatFrame，将需求ID 写成 "[ISSUE:"  需求ID "]"的格式
    :param Df_PTC_Spec:从Test Specification 中获取出来的DatFrame
    :param SpecCB: 导出的CodeBeamer的TestSpecification
    :return:df_SpecCB CodeBeamer的TestSpecification 的DataFrame
    """
    # ColumnsList = ["ID", "Priority","Name","Description","Pre-Action","Post-Action","Test Steps.Action","Test Steps.Expected result","Test Steps.Critical","Test Parameters","Verifies","Status","Type"]
    # ColumnsList = ["ID", "Parent", "Priority", "Name", "Description", "Pre-Action", "Post-Action",
    #                "Test Steps.Action", "Test Steps.Expected result",
    #                "Test Steps.Critical", "Test Parameters", "Verifies",
    #                "Status","Release", "Type", "_Original_TestCaseId", "_Test_Technique",
    #                "_VerifiesNonCbRequirements", "_LastReviewDate", "_ScriptPath",
    #                "_TestType", "_RegressionStrategy", "_FunctionGroup", "_Feature",
    #                "Functional Safety Relevant", "Cyber Security Relevant","Reserve_1","Test Method"]
    # 缺少一些非必要元素也是可以上传的
    ColumnsList = ["ID", "Parent", "Name",  "Verifies",
                   "Status","Test Method", "Release",
                   "_VerifiesNonCbRequirements"
                   ]
    df_spec_tocb = pd.DataFrame(columns= ColumnsList)

    print("GenerateSpec_CB_Init")
    df_spec_tocb["Name"] = "    " + Df_PTC_Spec["Name"]

    # df_spec_tocb["Status"] = Df_PTC_Spec["Status"].apply(ConverSpecTestResult2CBCaseStatus)
    df_spec_tocb["Status"] = "INIT"
    df_spec_tocb["Verifies"] = Df_PTC_Spec["Verifies"]
    df_spec_tocb["Release"] = CodeBeamer_Obj.Release
    df_spec_tocb["_VerifiesNonCbRequirements"] = Df_PTC_Spec["_VerifiesNonCbRequirements"]
    df_spec_tocb["Test Method"] = Df_PTC_Spec["Test Method"]

    with pd.ExcelWriter(SpecCB) as writer:
        To_Excel_Adjust_Weight(df_spec_tocb, writer, f'Export')
    # df_spec_tocb.to_excel(SpecCB, sheet_name="Export", index=False)
    DeleteLastEmptyRow(SpecCB)
    return df_spec_tocb

def DeleteLastEmptyRow(Spec):
    # 未知的问题，如果python生成了CB报告，但是无法上传，必须删除最后一行
    print("DeleteLastEmptyRow")
    wb = openpyxl.load_workbook(Spec)
    ws = wb.active
    max_row = ws.max_row
    ws.delete_rows(max_row + 1)
    wb.save(Spec)





def ReadSpecCB_FromCB(Spec_FromCB):
    """
    解析从CodeBeamer 下载下来的TestSpec,主要是为了知道已经存在的test case 的在CodeBeamer中的Case ID
    :param Spec_FromCB: 从CodeBeamer 下载下来的TestSpec
    :return:Df_SpecCB    从CodeBeamer 下载下来的TestSpec 的DataFrame
            Df_ID_Case  "ID", "Parent", "Name" 形成的DataFrame
    """
    print("&"*30 + "ReadSpecCB_FromCB2" + "&"*30)
    Df_SpecFromCB = pd.read_excel(Spec_FromCB, "Export", dtype=object)
    # print(Df_SpecCB.head(10))
    Df_SpecFromCB[["ID","Parent","Name"]] = Df_SpecFromCB[["ID","Parent","Name"]] .fillna(method = 'ffill')
    # print(Df_SpecCB.head(10))
    #获取版本号和issue ID
    Df_SpecFromCB= Df_SpecFromCB.fillna({"Release":"","Incident ID":"","Status":"","Verifies":""})
    Df_SpecFromCB["Release"] = Df_SpecFromCB["Release"].apply(GetReleaseValue)
    Df_SpecFromCB["Incident ID"] = Df_SpecFromCB["Incident ID"].apply(GetIncidentValue)
    Df_SpecFromCB["Verifies"] = Df_SpecFromCB["Verifies"].apply(GetIncidentValue)
    Df_SpecFromCB =  Df_SpecFromCB[["ID","Parent","Name","Verifies","Status","Type","Release","Incident ID"]]
    # Df_SpecCB.drop(0, inplace=True)


    Df_ID_Case = Df_SpecFromCB[["ID", "Parent", "Name","Type","Status"]].drop_duplicates(subset=['ID'])
    Df_ID_Case.drop(0, inplace=True)
    Df_ID_Case["Name"] = Df_ID_Case["Name"].str.strip()
    Df_ID_Case.set_index("Name",inplace = True,drop = False)
    Df_ID_Case = Df_ID_Case.astype({"ID": "int64", "Parent": "int64"})
    # Df_SpecCB = Df_SpecCB[(Df_SpecCB["Incident ID"] !="") | (Df_SpecCB["Release"] !="")]
    Df_SpecFromCB = Df_SpecFromCB[Df_SpecFromCB["Status"] != "Obsolete"] #删除status 部位obsolete的case
    # print(Df_ID_Case)
    # print(Df_ID_Case.columns)

    Df_SpecFromCB.drop(0, inplace=True)
    # print(Df_SpecCB.iloc[0, 2])
    return Df_SpecFromCB,Df_ID_Case



def GenerateSpec_CB_Modify(df_SpecCB_Generate,Df_ID_Case_FromCB,Df_SpecCB_FromCB,CodeBeamer_Obj,SpecCB_Modify):

    """
    解析从CodeBeamer 下载下来的TestSpec,
    1. 知道已经存在的test case 的在CodeBeamer中的Case ID
    2. 2022/8/23 修复informaiton的状态为空导致无法上传的bug

    :param df_SpecCB_Generate: 通过本工具生成的需要上传的Test Specification
    :param Df_ID_Case_FromCB:从CodeBeamer 下载下来的Test Specification
    :param Release 该测试case修改的版本号，CB版本号
    :param SpecCB_Modify: 修改后的Specification名称
    :return:NONE
    """
    print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&GenerateSpec_CB_Modify2&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
    df_SpecCB_Generate["Parent"] = Df_ID_Case_FromCB.iloc[1,1]
    #根据CodeBeamer Spec的ID给对应的Case赋值
    print("assign df_SpecCB_Generate Parent")

    print("#"*20)
    # print(Df_ID_Case_FromCB.index)

    df_SpecCB_Generate["ID"] = df_SpecCB_Generate["Name"].apply(GetCBID,df = Df_ID_Case_FromCB)


    print("assign df_SpecCB_Generate ID")

    #对不在最新的test specification的case ，设置属性为Obsolete
    # print(Df_ID_Case_FromCB)
    for caseName in Df_ID_Case_FromCB.index:
        print(caseName)
        # 如果 属性为information,则跳过，继续判断
        print(Df_ID_Case_FromCB.loc[caseName,"Type"])

        if Df_ID_Case_FromCB.loc[caseName,"Type"] == "Information":
            # print("2")
            continue;
        if caseName not in df_SpecCB_Generate["Name"].str.strip().values:
            print(caseName)
            Df_ID_Case_FromCB.loc[caseName,"Status"] = "Obsolete"
            Df_ID_Case_FromCB.loc[caseName,"Name"] = "    " + caseName
            df_SpecCB_Generate = df_SpecCB_Generate.append(Df_ID_Case_FromCB.loc[caseName])

            #从Df_SpecCB_FromCB 中删除掉这个case
            Df_SpecCB_FromCB = Df_SpecCB_FromCB[Df_SpecCB_FromCB["Name"].str.strip() != caseName]

    df_SpecCB_Generate["Release"] = CodeBeamer_Obj.Release


    with pd.ExcelWriter(SpecCB_Modify) as writer:
        To_Excel_Adjust_Weight(df_SpecCB_Generate, writer, f'Export')

    DeleteLastEmptyRow(SpecCB_Modify)
    # df_SpecCB_Generate.to_excel(SpecCB_Modify, sheet_name="Export", index=False)
    # DeleteLastEmptyRow(SpecCB_Modify)

def GetInitCaseList(FinalSpec):
    df = pd.read_excel(FinalSpec)

    InitCaseList = df.loc[df["ID"].isnull()]["Name"].drop_duplicates().str.strip().values

    print(type(InitCaseList) )
    print(InitCaseList)
    return InitCaseList

def To_Excel_Adjust_Weight(df: pd.DataFrame, writer: pd.ExcelWriter, sheet_name):
    """DataFrame保存为excel并自动设置列宽"""
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    #  计算表头的字符宽度
    # column_widths = (
    #     df.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
    # )
    column_widths = (
        df.columns.to_series().apply(lambda x: 15).values
    )
    #  计算每列的最大字符宽度
    # max_widths = (
    #     df.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
    # )
    max_widths = (
        df.astype(str).applymap(lambda x: 15).agg(max).values
    )

    # 计算整体最大宽度
    widths = np.max([column_widths, max_widths], axis=0)
    # 设置列宽
    worksheet = writer.sheets[sheet_name]
    worksheet.alignment = Alignment(wrap_text=True)
    # font = Font(color="FF0000")
    for i, width in enumerate(widths, 1):
        # openpyxl引擎设置字符宽度时会缩水0.5左右个字符，所以干脆+2使左右都空出一个字宽。
        worksheet.column_dimensions[get_column_letter(i)].width = width + 2
    # print(worksheet.max_row)
    for i in range(1,worksheet.max_row):
        worksheet.cell(column = 12,row = i).alignment = Alignment(wrap_text=True)


if __name__ == '__main__':
    pass
    Spec = r"C:\Users\victor.yang\Desktop\Work\CB\SpecTemplate\CHT_SWV_Project_FunctionName_Test Specification_Template_SC3.xlsm"
    # Df_PTC_Spec,CaseTrackerID,CB_Spec_Folder_ID,Release,TestRun_TrackerName= ReadSpec_TableOfContent(Spec)
    SpecCB = r"C:\Users\victor.yang\Desktop\Work\CB\SpecTemplate\SpecToCB.xlsx"
    SpecCB_FromCB = r"C:\Users\victor.yang\Desktop\Work\CB\84194_GWM_D30_RCS_SC2_2S - TC_L30_DES_SW_Test_Cases (4).xlsx"
    CB_Spec_Generate = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GMW_D30_2S_DCS_Test_Result_CodeBeamer.xlsx"
    CodeBeamer_Obj = CodeBeamer()
    print("Before" + CodeBeamer_Obj.TestRun_TrackerName)
    Df_Spec = ReadSpec_TableOfContent_SC3(Spec,CodeBeamer_Obj)
    print("After" + CodeBeamer_Obj.TestRun_TrackerName)
    GenerateSpec_CB_Init(Df_Spec, SpecCB,CodeBeamer_Obj)
    # ReadSpec_TableOfContent_0925(Spec)
    # GenerateSpec_CB_Init(Df_PTC_Spec,Release,CB_Spec_Generate)
    # df_SpecCB_FromCB, Df_ID_Case_FromCB =  ReadSpecCB_FromCB2(Spec_FromCB)
    # df_SpecCB_Generate = pd.read_excel(CB_Spec_Generate, "Export")
    # SpecCB_Modify = r"C:\Users\victor.yang\Desktop\Work\CB\CHT_SWV_GMW_D30_2S_DCS_Test_Result_CodeBeamer_Modify.xlsx"
    #
    #
    # GenerateSpec_CB_Modify2(df_SpecCB_Generate, Df_ID_Case_FromCB, df_SpecCB_FromCB,Release,
    #                                 SpecCB_Modify)
