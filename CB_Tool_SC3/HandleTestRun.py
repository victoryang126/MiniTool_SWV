import numpy as np
import pandas as pd
import openpyxl
import win32com
from win32com.client import Dispatch
import os
from CB_Tool_SC3.CodeBeamer import CodeBeamer
CaseResult = {

"Test Case DCS_1 - G_BBSD_DCS4_State":"PASSED",
"Test Case DCS_2 - G_BBSP_DCS6_State":"FAILED",
"Test Case DCS_3 - G_BB2L_DCS8_State":"PASSED",
"Test Case DCS_4 - G_BB2M_DCS2_State":"PASSED",
"Test Case DCS_5 - G_BB2R_DCS3_State":"PASSED",
"Test Case DCS_6 - G_SBRP_DCS5_State":"PASSED",
"Test Case DCS_7 - G_SBR2L_DCS7_State":"PASSED",
"Test Case DCS_8 - G_SBR2M_DCS13_State":"PASSED",
"Test Case DCS_9 - G_SBR2R_DCS1_State":"PASSED",
"Test Case DCS_10 - G_STSD_DCS10_State":"PASSED",
"Test Case DCS_11 - G_STSP_DCS9_State":"BLOCKED"
}

def ConvertExcelResult2TRunResult(result):
    """

    :param result: the cell value of _VerificationStatus in TableOfContent
    :return: return the corresonding value in CB
    """
    if result.strip().upper() == "OK":
        return "PASSED"
    elif result.strip().upper() == 'NOK':
        return "FAILED"
    else:
        return "NOT RUN YET"

def ReadResult_TableOfContent(Result,CodeBeamer_Obj):
    """
    读取Test specification 的TableOfContent
    然后将一个 case 对应一个单元格里面需求ID
    根据单元格的需求ID的数量,
    转换成多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    :param Result: Test specification
    :return:df_result 多行case（名字相同） 对应多行单元格的需求ID 的DataFrame
    """
    print("ReadResult_TableOfContent")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('max_colwidth', 200)
    # print(1)
    df_result = pd.read_excel(Result, "Table Of Contents", dtype='str')
    # print(2)
    ColumnsList = ["Object Text","_VerificationStatus","_Comment"]
    # if np.isnan(df_result.iloc[0,4]) else df_result.iloc[0,4].strip()
    #需要处理数据里面的前后的空格或者换行等特殊字符，

    CodeBeamer_Obj.Result_Summary= df_result.iloc[0, 4]
    CodeBeamer_Obj.TestRun_TrackerName = df_result.iloc[1, 4]
    CodeBeamer_Obj.CaseTrackerID = df_result.iloc[2, 4]
    CodeBeamer_Obj.CaseFolderID = df_result.iloc[3, 4]
    CodeBeamer_Obj.Release = df_result.iloc[4,4]

    Incident_ID_List = df_result["_Comment"].fillna("").values.tolist()
    CodeBeamer_Obj.Incident_IDs = ",".join(list(set([x for x in Incident_ID_List if x != ""])))


    # print(df_result)
    print("#"*30)
    # print("*"*30)
    # print(Result_Summary)
    # print(CaseTrackerID,CB_Spec_Folder_ID,Release)
    # print("*" * 30)
    df_result = df_result[ColumnsList]
    df_result = df_result.iloc[7:,:]
    df_result.columns = ["Name","RUN RESULT","Incident ID"]
    df_result["Name"] = df_result["Name"].str.strip() # 处理掉字符串前后的空格
    df_result.set_index("Name",inplace = True,drop = False)

    # print(df_result)
    df_result = df_result.fillna("undefined")
    df_result["RUN RESULT"] = df_result["RUN RESULT"].apply(ConvertExcelResult2TRunResult)
    # print(df_result)
    return df_result




def Handle_TestRun_Report(Excel,Df_Result):
    # df = pd.read_excel(Excel)
    # df.to_excel(Excel_Modify,"Run in Excel")

    print("#################Handle_TestRun_Report################################")
    wb = openpyxl.load_workbook(Excel)
    ws = wb.active
    # print(ws.max_row)
    #Name Row 3, col 2
    name_col =2
    # Run Result row 3, col 9
    result_col = 9
    # Release row3, Col 15
    release_col = 15
    start_row = 4 #所有数据从第四行开始,
    step_up = 3  #每3 行一个case

    #Test Run ID在Row 3, col 17
    # sheet.column_dimensions['Q'].hidden = 0
    #
    """
     1. Test Run 里面有相关case，但是Excel 里面没有，则报异常
     2. Test Run 里面没有相关case，但是Excel 里面有，则不处理。在update test case的时候工具会处理
    """
    # print(Df_Result.index)
    while start_row < ws.max_row:
        case_name = ws.cell(start_row,name_col).value
        # print(case_name)
        # print(Df_Result.index)
        if case_name in Df_Result.index:
            # print(case_name)
            ws.cell(start_row, result_col).value = Df_Result.loc[case_name,"RUN RESULT"]
            # ws.cell(start_row, release_col).value = Release
            start_row += step_up
        else:
            #2022/09/27 不处理这种异常，跳到下一个case
            # raise Exception(case_name +" not in index")
            print(case_name + " not in Table Of content")
            print(Df_Result.index)
            start_row += step_up
    wb.save(Excel)

    # 必须使用excel app 重新打开保存一下文件，否则会报错

    ExcelAPP = win32com.client.DispatchEx('Excel.Application')
    ExcelAPP.Visible = 0
    ExcelAPP.DisplayAlerts = 0
    wb = ExcelAPP.Workbooks.Open(os.path.abspath(Excel))
    wb.SaveAs(os.path.abspath(Excel))
    wb.Close(SaveChanges=0)
    ExcelAPP.Quit()
    print("#################Handle_TestRun_Report Finished################################")


if __name__ == "__main__":
    Excel = r"C:/Users/victor.yang/Desktop/Work/CB/SpecTemplate/VictorTest_2022_11_07 (1).xlsx"
    Excel_Modify = r"C:\Users\victor.yang\Downloads\Quick Test Run for 11 Test Cases at Sep 20 2022_Modify.xlsx"

    Result =r"C:/Users/victor.yang/Desktop/Work/CB/SpecTemplate/CHT_SWV_Project_FunctionName_Test Specification_Template_SC3.xlsm"
    CodeBeamer_Obj = CodeBeamer()
    Df_Result  = ReadResult_TableOfContent(Result,CodeBeamer_Obj)
    
    # Handle_TestRun_Report(Excel,Df_Result)